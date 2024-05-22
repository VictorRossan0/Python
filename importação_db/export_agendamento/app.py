import os
import csv
import pandas as pd
from flask import Flask, send_file, request, render_template, send_from_directory, redirect, url_for, flash
from sqlalchemy import create_engine
from datetime import datetime
from openpyxl import load_workbook
from io import StringIO
from werkzeug.utils import secure_filename

# Obter o caminho absoluto do diretório 'uploads' em relação ao diretório atual do script
uploads_dir = os.path.abspath('uploads')

# Verificar se o diretório 'uploads' existe e criar se não existir
if not os.path.exists(uploads_dir):
    os.makedirs(uploads_dir)

app = Flask(__name__)
app.secret_key = '72UJ7qyr1hAps70J3Y3nU7ovH0D0CMqH'
app.config['UPLOAD_FOLDER'] = uploads_dir  # Diretório onde os arquivos serão salvos

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/export')
def export_to_excel():
    # Obter a data e hora atuais
    now = datetime.now()
    timestamp = now.strftime("%d-%m_%H-%M") # Formato DD-MM_HH:MM
    
    # Definir o nome do arquivo com a data e hora
    filename = f'Agendamento_{timestamp}.xlsx'
    
    # Definir a query
    query = 'SELECT cir, oi, cm, regional, cliente, macro, bloco, gross, pendencia, status, tarefa_inter, tarefa_ant, analista, designacao, terminal, servico, tipo_alt, tipo_serv, projeto, aging_macro_processo_sem_inter, aging_tarefa, aging_status_disponivel, inter, log, cli_black, cli_difer, ativ_agend, obs_agend, data_pre_agend, data_solu_pend, data_agend FROM jarvas.jarvas_agendamento_13 WHERE ativo = 1'
    
    # Conectar ao banco de dados, executar a query e criar um DataFrame com o resultado
    df = pd.read_sql_query(query, create_engine('mysql+pymysql://victor:victor@10.230.43.179:3306/jarvas'))

    # Renomear as colunas do DataFrame
    df.columns = ['Código Cir', 'OI', 'CM', 'Regional', 'Cliente', 'Macro Processo', 'Bloco', 'Gross', 'Tarefa', 'Status', 'Tarefa Interrupção', 'Tarefa Anterior', 'User Tratamento', 'Designação', 'Terminal', 'Serviço', 'Tipo Alteração', 'Tipo Serviço', 'Projeto', 'Aging Macro Processo sem Interrupção', 'Aging Tarefa', 'Aging Status Disponível', 'Interrupções', 'Log', 'Cli. Blacklist', 'Cli. Diferenciado', 'Ativ. Agendamento', 'Obs Agendamento', 'Data Pré Agendamento', 'Data Solu. Pend.', 'Data Agendamento']

    # Exportar o DataFrame
    df.to_excel(os.path.join(app.config['UPLOAD_FOLDER'], filename), index=False)
    print(f"Dados exportados com sucesso para o arquivo {filename}")
    
    # Retornar o arquivo para download
    return send_file(os.path.join(app.config['UPLOAD_FOLDER'], filename), as_attachment=True)

@app.route('/exportar-wf')
def exportar_usuarios():
    
    # Definir o nome do arquivo com a data e hora
    filename = f'cir_wf.xlsx'

    # Definir a query
    query = """
    SELECT id, cir, regional, cm, tecnologia, servico, atividade_remota, designacao, pcl, obs, velocidade, oe, created_at
    FROM consulta_cir
    WHERE obs IS NULL OR obs = ''
    """

    # Conectar ao banco de dados, executar a query e criar um DataFrame com o resultado
    engine = create_engine('mysql+pymysql://victor:victor@10.230.43.179:3306/jarvas')
    df = pd.read_sql_query(query, engine)

    # Exportar o DataFrame
    df.to_excel(os.path.join(app.config['UPLOAD_FOLDER'], filename), index=False)
    print(f"Dados exportados com sucesso para o arquivo {filename}")

    # Retornar o arquivo para download
    return send_file(os.path.join(app.config['UPLOAD_FOLDER'], filename), as_attachment=True)

@app.route('/import', methods=['POST'])
def import_from_csv():
    # Verificar se foi enviado um arquivo
    if 'file' not in request.files:
        return 'Nenhum arquivo enviado'

    file = request.files['file']
    if file.filename == '':
        return 'Nenhum arquivo selecionado'

    # Ler o arquivo CSV com a codificação correta (utf-8)
    df = pd.read_csv(file, encoding='utf-8')

    # Iterar sobre as colunas do DataFrame e converter as datas para o formato correto
    for col in df.columns:
        if df[col].dtype == 'datetime64[ns]':
            df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%Y-%m-%d %H:%M:%S')

    # Conectar ao banco de dados e inserir os dados
    engine = create_engine('mysql+pymysql://victor:victor@10.230.43.179:3306/jarvas')
    df.to_sql('produtividade_agendamento', con=engine, if_exists='append', index=False)

    flash(f'Dados do arquivo {file.filename} importados com sucesso', 'success')
    return redirect(url_for('index'))

@app.route('/upload_excel', methods=['POST'])
def upload_excel():
    # Verificar se foi enviado um arquivo
    if 'excel_file' not in request.files:
        return 'Nenhum arquivo enviado'

    excel_file = request.files['excel_file']
    if excel_file.filename == '':
        return 'Nenhum arquivo selecionado'

    # Salvar o arquivo Excel temporariamente
    excel_filename = secure_filename(excel_file.filename)
    excel_file.save(os.path.join(app.config['UPLOAD_FOLDER'], excel_filename))

    # Converter o arquivo Excel para CSV
    csv_filename = convert_excel_to_csv(os.path.join(app.config['UPLOAD_FOLDER'], excel_filename))
    if csv_filename is None:
        return 'Erro ao converter arquivo Excel para CSV'

    # Retornar o arquivo CSV para download
    return send_file(os.path.join(app.config['UPLOAD_FOLDER'], csv_filename), as_attachment=True)


def convert_excel_to_csv(excel_filename):
    csv_filename = excel_filename.replace('.xlsx', '.csv')
    try:
        wb = load_workbook(excel_filename, read_only=True)
        sheet = wb.active
    except FileNotFoundError:
        return None  # Retorna None se o arquivo não for encontrado

    csv_data = StringIO()
    csv_writer = csv.writer(csv_data)
    for row in sheet.iter_rows(values_only=True):
        converted_row = []
        for value in row:
            if isinstance(value, datetime):
                # Verifica se o valor possui segundos
                if len(value.strftime('%S')) == 2:
                    converted_row.append(value.strftime('%Y-%m-%d %H:%M:%S'))
                else:
                    converted_row.append(value.strftime('%Y-%m-%d %H:%M'))
            else:
                converted_row.append(value)
        csv_writer.writerow(converted_row)

    with open(csv_filename, 'w', encoding='utf-8') as csv_file:
        csv_file.write(csv_data.getvalue())

    return csv_filename


@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

if __name__ == '__main__':
    app.run(host='0.0.0.0', debug=True)
