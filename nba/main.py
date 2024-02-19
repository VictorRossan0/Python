from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter
from Models.calendario_nba import extrair_info_calendario
from Models.leaders_nba import extrair_info_lideres
from Models.teams_nba_leaders import extrair_info_teams
from Models.doubled_tripled_nba import main as doubled_tripled_main

# Extrai as informações uma vez
extrair_info_calendario()
extrair_info_lideres()
extrair_info_teams()
doubled_tripled_main()

def criar_novo_excel():
    wb_novo = Workbook()
    wb_novo.remove(wb_novo.active)  # Remove a folha em branco padrão

    # Abre os arquivos Excel existentes
    wb_calendario = load_workbook('Excel/informacoes_eventos.xlsx')
    wb_lideres = load_workbook('Excel/leaders_nba.xlsx')
    wb_teams = load_workbook('Excel/team_leaders_nba.xlsx')
    wb_dd = load_workbook('Excel/doubledouble.xlsx')
    wb_td = load_workbook('Excel/tripledouble.xlsx')

    # Copia as abas para o novo arquivo Excel
    for sheet in wb_calendario.sheetnames:
        ws = wb_calendario[sheet]
        ws_novo = wb_novo.create_sheet(title=sheet)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            values = [cell.value for cell in row]
            ws_novo.append(values)

    for sheet in wb_lideres.sheetnames:
        ws = wb_lideres[sheet]
        ws_novo = wb_novo.create_sheet(title=sheet)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            values = [cell.value for cell in row]
            ws_novo.append(values)
    
    for sheet in wb_teams.sheetnames:
        ws = wb_teams[sheet]
        ws_novo = wb_novo.create_sheet(title=sheet)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            values = [cell.value for cell in row]
            ws_novo.append(values)

    for sheet in wb_dd.sheetnames:
        ws = wb_dd[sheet]
        ws_novo = wb_novo.create_sheet(title=sheet)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            values = [cell.value for cell in row]
            ws_novo.append(values)
    
    for sheet in wb_td.sheetnames:
        ws = wb_td[sheet]
        ws_novo = wb_novo.create_sheet(title=sheet)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            values = [cell.value for cell in row]
            ws_novo.append(values)

    # Ajuste automático da largura da coluna
    for sheet in wb_novo.sheetnames:
        ws = wb_novo[sheet]
        for col in ws.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            
            # Converta o número da coluna em uma string
            column_letter = get_column_letter(column)
            
            ws.column_dimensions[column_letter].width = adjusted_width

    # Salva o novo arquivo Excel
    wb_novo.save('Excel/planilha_NBA.xlsx')

if __name__ == "__main__":
    criar_novo_excel()
    print(f"\nTodas as informações salvas em planilha_NBA.xlsx")
