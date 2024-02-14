import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.firefox.options import Options

def extrair_dados_tabela():
    print("Abrindo o navegador")
    firefox_options = Options()
    firefox_options.headless = True
    driver = webdriver.Firefox(options=firefox_options)

    url = 'https://redscores.com/pt-br/prognosticos/estatisticas-de-futbol/teams'
    driver.get(url)
    print("Navegador aberto")

    # Encontrar a tabela dentro da div com a classe "league"
    tabela_xpath = '//div[@class="league"]//table[@class="table-data__table"]'
    tabela_element = driver.find_element('xpath', tabela_xpath)

    # Extrair os dados da tabela
    linhas = tabela_element.find_elements('xpath', './/tbody/tr')
    
    with open('TXT/dados_extraidos.txt', 'w', encoding='utf-8') as arquivo:
        # Escrever os cabeçalhos das colunas
        arquivo.write("Ligas\tProgresso\tMédia de Gols\t+0.5\t+1.5\t+2.5\t+3.5\t+4.5\tBTTS\tMédia de Chutes\tMédia de Chutes no Alvo\tMédia de Escanteios\tMédia de Cartões\tAtaques Perigosos\n")
        
        # Escrever os dados das linhas
        equipe_liga_anterior = None
        for linha in linhas:
            # Extrair o nome da equipe e da liga
            equipe_liga = linha.find_element('xpath', './/td[@class="sticky-left left01"]').text.strip()
            
            # Verificar se a equipe e liga atual é igual à anterior
            if equipe_liga == equipe_liga_anterior:
                dados_linha[0] = ''  # Se for igual, deixamos a primeira célula vazia
            else:
                dados_linha = [equipe_liga]  # Caso contrário, armazenamos a nova equipe e liga
                
            # Extrair os demais dados da linha
            colunas = linha.find_elements('tag name', 'td')
            dados_linha += [coluna.text for coluna in colunas[1:]]
            
            arquivo.write('\t'.join(dados_linha) + '\n')
            
            equipe_liga_anterior = equipe_liga
            
    print("Dados extraídos da tabela e salvos no arquivo 'dados_extraidos.txt'.")
    
    driver.quit()

def criar_excel(colunas_selecionadas=None):
    # Lendo os dados do arquivo TXT
    dados = pd.read_csv('TXT/dados_extraidos.txt', sep='\t')
    
    # Selecionando apenas as colunas desejadas, se especificadas
    if colunas_selecionadas:
        dados = dados[colunas_selecionadas]

    # Criando um novo arquivo Excel
    wb = Workbook()
    planilha = wb.active
    planilha.title = "Estatísticas TOP5 Ligas"

    # Escrever os cabeçalhos na planilha
    cabecalhos = dados.columns.tolist()
    planilha.append(cabecalhos)

    # Escrever os dados na planilha
    for _, row in dados.iterrows():
        planilha.append(row.tolist())

    # Ajuste automático da largura das colunas
    for coluna in planilha.columns:
        comprimento_maximo = 0
        for valor in coluna:
            comprimento = len(str(valor))
            if comprimento > comprimento_maximo:
                comprimento_maximo = comprimento
        coluna_letra = get_column_letter(coluna[0].column)
        planilha.column_dimensions[coluna_letra].width = comprimento_maximo + 2

    # Salvando o arquivo Excel
    wb.save('Excel/dados_extraidos.xlsx')

    print("Arquivo Excel criado com sucesso.")

# Chamando as funções
extrair_dados_tabela()

# Escolha as colunas que deseja salvar no Excel (substitua None pelas colunas desejadas)
colunas_selecionadas = ["Ligas", "Média de Gols", "+0.5", "+1.5", "+2.5", "Chutes", "Escanteios", "Cartões"]
criar_excel(colunas_selecionadas)
