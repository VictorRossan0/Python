from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter
from Models.eredivisie import eredivisie
from Models.premier_russia import premier_russia
from Models.primeira_liga import primeira_liga
from Models.super_lig import super_lig
from Models.brasileirao import brasileirao


def copiar_abas(wb_origem, wb_destino):
    for sheet in wb_origem.sheetnames:
        ws = wb_origem[sheet]
        ws_novo = wb_destino.create_sheet(title=sheet)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            values = [cell.value for cell in row]
            ws_novo.append(values)

def ajustar_largura(wb):
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_length = 0
            column = col[0].column
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            column_letter = get_column_letter(column)
            ws.column_dimensions[column_letter].width = max_length + 2

def criar_novo_excel():
    wb_novo = Workbook()
    wb_novo.remove(wb_novo.active)  # Remove a folha em branco padrão

    # Chama as funções para extrair informações
    eredivisie()
    premier_russia()
    primeira_liga()
    super_lig()
    brasileirao()

    arquivos_ligas = [
        ('Excel/dados_bundesliga.xlsx', 'Bundesliga'),
        ('Excel/dados_eredivisie.xlsx', 'Eredivisie'),
        ('Excel/dados_laliga.xlsx', 'La Liga'),
        ('Excel/dados_ligue1.xlsx', 'Ligue 1'),
        ('Excel/dados_premier.xlsx', 'Premier League'),
        ('Excel/dados_premier_russia.xlsx', 'Premier Russia'),
        ('Excel/dados_primeira_liga.xlsx', 'Primeira Liga'),
        ('Excel/dados_serieA.xlsx', 'Serie A'),
        ('Excel/dados_super_lig.xlsx', 'Super Lig'),
        ('Excel/dados_tabela_brasileirao.xlsx', 'Brasileirão')
    ]

    for arquivo, nome in arquivos_ligas:
        try:
            wb_liga = load_workbook(arquivo)
            copiar_abas(wb_liga, wb_novo)
        except Exception as e:
            print(f"Erro ao abrir ou copiar {nome}: {e}")

    ajustar_largura(wb_novo)
    wb_novo.save('Excel/planilha_TOP10_ligas.xlsx')

if __name__ == "__main__":
    criar_novo_excel()
    print(f"\nTodas as informações salvas em planilha_TOP10_ligas.xlsx")